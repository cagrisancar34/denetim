from flask import Flask, render_template, request, jsonify, send_file
from data_utils import load_data, save_data
from score_utils import calculate_score, get_failed_questions, get_category_scores, extract_answers
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import os
import json
import pandas as pd

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///denetimler.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

class Denetim(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tesis_adi = db.Column(db.String(128))
    tarih = db.Column(db.String(32))
    denetim_yapan = db.Column(db.String(128))
    cevaplar = db.Column(db.Text)  # JSON string olarak saklanacak
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

DATA_FILE = os.path.join(app.root_path, "data.json")
categories = load_data()

@app.before_request
def create_tables():
    db.create_all()

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/add_category", methods=["POST"])
def add_category():
    category = request.form.get("category")
    if category and category not in categories:
        categories[category] = []
        save_data(categories)
        return jsonify({"success": True, "categories": list(categories.keys())})
    return jsonify({"success": False})

@app.route("/add_question", methods=["POST"])
def add_question():
    category = request.form.get("category")
    question = request.form.get("question")
    if category in categories and question:
        categories[category].append(question)
        save_data(categories)
        return jsonify({"success": True, "questions": categories[category]})
    return jsonify({"success": False})

@app.route("/add_questions", methods=["POST"])
def add_questions():
    data = request.get_json()
    category = data.get("category")
    questions = data.get("questions", [])
    if category in categories and questions:
        for q in questions:
            if q:
                categories[category].append(q)
        save_data(categories)
        return jsonify({"success": True, "questions": categories[category]})
    return jsonify({"success": False})

@app.route("/submit", methods=["POST"])
def submit():
    form_data = request.form.to_dict()
    answers = extract_answers(form_data, categories)
    score, total = calculate_score(answers, categories)
    failed_questions = get_failed_questions(answers, categories)
    category_scores = get_category_scores(answers, categories)
    denetim = Denetim(
        tesis_adi=form_data.get("tesis_adi", ""),
        tarih=form_data.get("denetim_tarihi", ""),
        denetim_yapan=form_data.get("denetim_yapan", ""),
        cevaplar=json.dumps(answers)
    )
    db.session.add(denetim)
    db.session.commit()
    return render_template("submit.html", form_data=answers, categories=categories, score=score, total=total, failed_questions=failed_questions, category_scores=category_scores)

@app.route("/kategori")
def kategori():
    return render_template("kategori.html", categories=categories)

@app.route("/tesis_karnesi")
def tesis_karnesi():
    # Son denetim kaydını veritabanından oku
    d = Denetim.query.order_by(Denetim.created_at.desc()).first()
    if d:
        last = json.loads(d.cevaplar)
    else:
        last = {}
    score, total = calculate_score(last, categories)
    failed_questions = get_failed_questions(last, categories)
    category_scores = get_category_scores(last, categories)
    return render_template("tesis_karnesi.html", form_data=last, categories=categories, score=score, total=total, failed_questions=failed_questions, category_scores=category_scores)

@app.route("/get_last_denetim")
def get_last_denetim():
    d = Denetim.query.order_by(Denetim.created_at.desc()).first()
    if d:
        cevaplar = json.loads(d.cevaplar)
        cevaplar["tesis_adi"] = d.tesis_adi
        cevaplar["tarih"] = d.tarih
        cevaplar["denetim_yapan"] = d.denetim_yapan
        return jsonify({"form_data": cevaplar, "categories": categories})
    else:
        return jsonify({"form_data": {}, "categories": categories})

@app.route("/get_all_denetimler")
def get_all_denetimler():
    denetimler = Denetim.query.order_by(Denetim.created_at).all()
    result = []
    for d in denetimler:
        cevaplar = json.loads(d.cevaplar)
        cevaplar["tesis_adi"] = d.tesis_adi
        cevaplar["tarih"] = d.tarih
        cevaplar["denetim_yapan"] = d.denetim_yapan
        result.append(cevaplar)
    return jsonify({"denetimler": result, "categories": categories})

@app.route("/tesis_karnesi_data")
def tesis_karnesi_data():
    tesis_adi = request.args.get("tesis_adi", "")
    denetimler = Denetim.query.order_by(Denetim.created_at).all()
    filtered = []
    for d in denetimler:
        cevaplar = json.loads(d.cevaplar)
        if not tesis_adi or d.tesis_adi == tesis_adi:
            cevaplar["tesis_adi"] = d.tesis_adi
            cevaplar["tarih"] = d.tarih
            cevaplar["denetim_yapan"] = d.denetim_yapan
            filtered.append(cevaplar)
    if filtered:
        form_data = filtered[-1]
        score, total = calculate_score(form_data, categories)
        category_scores = get_category_scores(form_data, categories)
        failed_questions = get_failed_questions(form_data, categories)
        return jsonify({
            "score": score,
            "total": total,
            "category_scores": category_scores,
            "failed_questions": failed_questions,
            "form_data": form_data
        })
    else:
        return jsonify({
            "score": 0,
            "total": 0,
            "category_scores": {},
            "failed_questions": [],
            "form_data": {}
        })

@app.route("/export_excel")
def export_excel():
    denetimler = Denetim.query.order_by(Denetim.created_at).all()
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    import tempfile

    if not denetimler:
        return "Hiç denetim kaydı yok.", 404

    # Son denetimi al
    d = denetimler[-1]
    cevaplar = json.loads(d.cevaplar)
    wb = Workbook()
    ws = wb.active
    ws.title = "Denetim"
    # 1. satır: Tesis Adı
    ws["A1"] = "Tesis Adı"
    ws["B1"] = cevaplar.get("tesis_adi", d.tesis_adi)
    # 2. satır: Denetim Tarihi
    ws["A2"] = "Denetim Tarihi"
    ws["B2"] = cevaplar.get("denetim_tarihi", d.tarih)
    # 3. satır: Denetim Yapan
    ws["A3"] = "Denetim Yapan"
    ws["B3"] = cevaplar.get("denetim_yapan", d.denetim_yapan)
    # 5. satır: başlıklar
    ws["A5"] = "Soru"
    ws["B5"] = "Yanıt"
    ws["C5"] = "Açıklama"
    ws["D5"] = "Kategori"
    bold_font = Font(bold=True)
    for col in range(1, 5):
        ws[f"{get_column_letter(col)}5"].font = bold_font
    row_idx = 6
    for category, questions in categories.items():
        for question in questions:
            yanit = cevaplar.get(question, "-")
            aciklama = cevaplar.get(f"aciklama_{question}", "-")
            ws[f"A{row_idx}"] = question
            ws[f"B{row_idx}"] = yanit
            ws[f"C{row_idx}"] = aciklama
            ws[f"D{row_idx}"] = category
            row_idx += 1
    # Dosyayı geçici olarak kaydet
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        return send_file(tmp.name, as_attachment=True, download_name="denetim_sonuclari.xlsx")

@app.route("/export_excel_page", methods=["GET"])
def export_excel_page():
    denetimler = Denetim.query.order_by(Denetim.created_at).all()
    tesisler = list({d.tesis_adi for d in denetimler if d.tesis_adi})
    return render_template("export_excel.html", tesisler=tesisler)

@app.route("/export_excel_filtered")
def export_excel_filtered():
    tesis_adi = request.args.get("tesis_adi", "")
    denetim_tarihi = request.args.get("denetim_tarihi", "")
    denetim_yapan = request.args.get("denetim_yapan", "")
    denetimler = Denetim.query.order_by(Denetim.created_at).all()
    # Sadece ilgili denetimi bul
    filtered = [d for d in denetimler if d.tesis_adi == tesis_adi and d.tarih == denetim_tarihi and d.denetim_yapan == denetim_yapan]
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    import tempfile
    if not filtered:
        return "Seçilen denetime ait kayıt yok.", 404
    d = filtered[0]
    cevaplar = json.loads(d.cevaplar)
    wb = Workbook()
    ws = wb.active
    ws.title = "Denetim"
    ws["A1"] = "Tesis Adı"
    ws["B1"] = d.tesis_adi
    ws["A2"] = "Denetim Tarihi"
    ws["B2"] = d.tarih
    ws["A3"] = "Denetim Yapan"
    ws["B3"] = d.denetim_yapan
    ws["A5"] = "Soru"
    ws["B5"] = "Yanıt"
    ws["C5"] = "Açıklama"
    ws["D5"] = "Kategori"
    bold_font = Font(bold=True)
    for col in range(1, 5):
        ws[f"{get_column_letter(col)}5"].font = bold_font
    row_idx = 6
    for category, questions in categories.items():
        for question in questions:
            yanit = cevaplar.get(question, "-")
            aciklama = cevaplar.get(f"aciklama_{question}", "-")
            ws[f"A{row_idx}"] = question
            ws[f"B{row_idx}"] = yanit
            ws[f"C{row_idx}"] = aciklama
            ws[f"D{row_idx}"] = category
            row_idx += 1
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        return send_file(tmp.name, as_attachment=True, download_name="denetim_sonuclari.xlsx")

@app.route("/form")
def form():
    return render_template("form.html", categories=categories)

if __name__ == "__main__":
    app.run(debug=True)
